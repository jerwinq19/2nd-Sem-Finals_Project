# int necessary lib 
from tkinter import *
from tkinter import ttk, simpledialog, filedialog
import json, os
from PIL import ImageTk, Image
from tkinter import messagebox
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from reportlab.lib.pagesizes import letter,landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from LIBS.tools import encrypt_pw
import customtkinter as ctk

def main_window(tst):
    def show_frames(frame):
        '''Dis play the frames when button is clicked for
        smooth navigating in software'''
        frame.tkraise()
    
    def updater(event):
        update_TV()
        filter_data()

    tst.withdraw()  
    root = Toplevel()
    root.title("Admin side")
    
    app_Width,app_Height = 1025, 750
    screen_width,screen_height = root.winfo_screenwidth(), root.winfo_screenheight()

    x = (screen_width / 2) - (app_Width / 2)
    y = (screen_height / 2) - (app_Height / 2)
    
    root.geometry(f"{app_Width}x{app_Height}+{int(x)}+{int(y)}")
    root.resizable(False,False)

    side_bar_frame = Frame(root,width=90, height=750,bg="#379777")
    side_bar_frame.place(x=0,y=0) # 
    side_font = ("Roboto", 13, "bold")

    # IMAGES
    db_img = Image.open("ASSETS\\dashboard.png")
    list_img = Image.open("ASSETS\\list.png")
    iv_img = Image.open("ASSETS\\inventory.png")
    sts_img = Image.open("ASSETS\\settings.png")
    brgr_img = Image.open("ASSETS\\burger_menu.png")


    dash_board = ctk.CTkImage(light_image=db_img, size=(50,50))
    list_icon = ctk.CTkImage(light_image=list_img, size=(50,50))
    inventory_icon = ctk.CTkImage(light_image=iv_img, size=(50,50))
    settings_icon = ctk.CTkImage(light_image=sts_img, size=(50,50))
    burger_menu = ctk.CTkImage(light_image=brgr_img, size=(50,50))

    new_order = ImageTk.PhotoImage(Image.open("ASSETS\\new_order.png").resize((25,25)))

    container = Frame(root, bg="#F5F7F8") # This it  the main frame the Software
    container.place(x=100, y=0, width=1200, height=750)
    
    # int necessary Frames
    dashboard_page = Frame(container, bg="#f2f2f2")
    orders_page = Frame(container, bg="#f2f2f2")
    inventory_page = Frame(container, bg="#f2f2f2")
    settings_page = Frame(container, bg="#f2f2f2")
    new_order_page = Frame(container, bg="#f2f2f2")
    about_us_frame = Frame(container, bg="#f2f2f2")
    # placing of frames
    dashboard_page.place(x=0,y=0, width=1200, height=750)
    orders_page.place(x=0,y=0, width=1200, height=750)
    inventory_page.place(x=0,y=0, width=1200, height=750)
    settings_page.place(x=0,y=0, width=1200, height=750)
    new_order_page.place(x=0, y=0, width=1200, height=750)

    # SETTINGS PAGES WIDGETS
    settings_pageName = Label(settings_page, text="Settings",font=("Serif", 32, "bold"), fg="#379777").place(x=30, y=30)
    
    change_file_path_lbl = Label(settings_page, text="Excel/PDF file path",font=("Serif", 20, "bold"),fg="#379777")
    change_file_path_lbl.place(x=30, y=100)
    
    path_frame = ctk.CTkFrame(settings_page, width=880, height=70, border_color="#000000", border_width=2, fg_color="#f2f2f2", corner_radius=10)
    path_frame.place(x=30, y=150)
    path_frame.pack_propagate(False)
    
    path_lbl = Label(path_frame, text="", font=("Serif", 25,"bold"))
    path_lbl.pack(side=LEFT, padx=10, pady=10)

    add_file_path = ImageTk.PhotoImage(Image.open("ASSETS\\add_file.png").resize((50,50)))
    select_path_btn = Button(path_frame, image=add_file_path, relief=FLAT, cursor="hand2", command=lambda:select_path())
    select_path_btn.pack(side=RIGHT, padx=5)
    
    def universal_path():
        # sets the initial directory for the files to save
        global user_home_path
        user_home_path = os.path.expanduser("~\\Documents")
        path_lbl.config(text=user_home_path)
        print(user_home_path)
            
    def select_path() -> str:
        # ask user to select specific path to save the reports
        try:
            file_name = filedialog.askdirectory(initialdir="d:\Documents", title="Select dir")
            path_lbl.config(text=file_name)
            if file_name == "":
                path_lbl.config(text=user_home_path)
            # if_there_file_checker()
        except FileNotFoundError:
            path_lbl.config(text=f"YOUR//PATH//GOES//HER")
            return None
    
    def burger_menu_animation():
        ''''Toggel and resize the root dimenson and as well configure the postion and width of the menu bar 
            to mimic the animation of a collapsible menu'''
        nonlocal is_on
        if is_on:
            # return to normal
            side_bar_frame.place_configure(width=250, height=750)
            # menu_name_lbl = Label(side_bar_frame, text="Shop-IT", font=("Serif", 25, "bold")).place(x=50, y=30)
            menu_bar.place_configure(x=180, y=20)
            
            dashB_lbl.place_configure(x=10, y=200)
            dashB_lbl.configure(text="Dashboard", width=200)
            
            orders_lbl.place_configure(x=10, y=275)
            orders_lbl.configure(text="Manage Orders", width=200)
            
            inventory_lbl.place_configure(x=10, y=350)
            inventory_lbl.configure(text="Inventory", width=200)
            
            settings_lbl.place_configure(x=10, y=420)
            settings_lbl.configure(text="Settings", width=200)
        
            orders_label.place_configure(x=30, y=30)
            container.place_configure(x=275, y=0)
            root.geometry("1200x750")

            is_on = False 
        else:
            # Collapsed
            side_bar_frame.place_configure(width=80, height=750)
            menu_bar.place_configure(x=10, y=20)
            
            dashB_lbl.place_configure(x=10, y=200)
            dashB_lbl.configure(text="", width=40)
            
            orders_lbl.place_configure(x=10, y=275)
            orders_lbl.configure(text="", width=40)
            
            inventory_lbl.place_configure(x=10, y=350)
            inventory_lbl.configure(text="", width=40)
            
            settings_lbl.place_configure(x=10, y=420)
            settings_lbl.configure(text="", width=40)
            
            container.place_configure(x=100, y=0)
            root.geometry("1025x750")
    
            is_on = True
            
    # side menu bar
    menu_bar = ctk.CTkButton(side_bar_frame, text="",width=50,image=burger_menu,command=lambda:burger_menu_animation(),cursor="hand2", border_color="#379777", fg_color="transparent", hover=False)
    menu_bar.place(x=12, y=20)

    # bool var for on and off
    is_on = True

    # Dashboard page
    dashB_lbl = ctk.CTkButton(side_bar_frame, text="", font=side_font, text_color="#f2f2f2",fg_color="transparent", width=40, cursor="hand2", image=dash_board,compound=LEFT,command=lambda: show_frames(dashboard_page), anchor=EW, hover_color="#333333")
    dashB_lbl.place(x=10, y=200)

    orders_lbl = ctk.CTkButton(side_bar_frame, text="", font=side_font, text_color="#f2f2f2",fg_color="transparent", width=40, cursor="hand2", image=list_icon, compound=LEFT,command=lambda: show_frames(orders_page), anchor=EW, hover_color="#333333")
    orders_lbl.place(x=10, y=275)

    inventory_lbl = ctk.CTkButton(side_bar_frame, text="", font=side_font, text_color="#f2f2f2",fg_color="transparent", width=40, cursor="hand2", image=inventory_icon, compound=LEFT,command=lambda: show_frames(inventory_page), anchor=EW, hover_color="#333333")
    inventory_lbl.place(x=10, y=350)

    settings_lbl = ctk.CTkButton(side_bar_frame, text="", font=side_font, text_color="#f2f2f2",fg_color="transparent", width=40, cursor="hand2", image=settings_icon, compound=LEFT,command=lambda: show_frames(settings_page), anchor=EW, hover_color="#333333")
    settings_lbl.place(x=10, y=420)
    
    # ------------------------------------------------------------------------------------------------------------------------------------------------------
    orders_label = Label(dashboard_page, text="Dash Board", font=("Serif", 32, "bold"),fg="#379777", bg="#f2f2f2")
    orders_label.place(x=30,y=30)

    charts_frames = ctk.CTkFrame(dashboard_page, width=880, height=350, fg_color="#f2f2f2", border_color="#333333", border_width=3)
    charts_frames.place(x=30, y=100)

    # total inventory items frame
    total_inventory_items = ctk.CTkFrame(dashboard_page, fg_color="#f2f2f2", border_color="#333333", border_width=3, width=280, height=250)
    total_inventory_items.place(x=30, y=470)

    total_item_label = Label(total_inventory_items, text="Total Inventory Item", font=("Serif", 14,"bold"), bg="#f2f2f2",fg="#379777")
    total_item_label.place(x=10,y=10)

    total_item_label_counter = Label(total_inventory_items, text="000", font=("Serif", 30, "bold"), bg="#f2f2f2", fg="#333333")
    total_item_label_counter.place(x=130, y=130, anchor=CENTER)

    # total sales frame 
    total_sales_frame = ctk.CTkFrame(dashboard_page, width=280, height=250, border_color="#333333", border_width=3, fg_color="#f2f2f2")
    total_sales_frame.place(x=330, y=470)

    total_sales_lbl = Label(total_sales_frame, text="Total sales", font=("Serif", 14,"bold"), bg="#f2f2f2",fg="#379777")
    total_sales_lbl.place(x=10,y=10)

    total_sales_counter_lbl = Label(total_sales_frame, text="000", font=("Serif", 30, "bold"), bg="#f2f2f2", fg="#333333")
    total_sales_counter_lbl.place(x=130, y=130, anchor=CENTER)

    # total revenue frame
    revenue_frame =ctk.CTkFrame(dashboard_page, width=280, height=250, border_color="#333333", border_width=3, fg_color="#f2f2f2")
    revenue_frame.place(x=630, y=470)

    revenue_lbl = Label(revenue_frame, text="Total Earnings", font=("Serif", 14,"bold"), bg="#f2f2f2",fg="#379777")
    revenue_lbl.place(x=10,y=10)

    revenue_lbl_counter = Label(revenue_frame, text="000", font=("Serif", 30, "bold"), bg="#f2f2f2", fg="#333333")
    revenue_lbl_counter.place(x=130, y=130, anchor=CENTER)


    def create_order_status():
        '''
        Function do? Make bar graph using Matplotlib
        by extracting the status and putting to the status_counter

        '''
        global bar_var, mycanvas
        status_counter = {}
        with open("CREDS/orders.json", 'r') as readF:
            read = json.load(readF)

        for each in read["inventory_orders"]:
            '''This extract the status'''
            status = each["Status"]
            if status in status_counter:
                '''increment 1 if the status exist in the status counter'''
                status_counter[status] += 1
            else:
                '''count the status'''
                status_counter[status] = 1
        
        sorted_status = sorted(status_counter.items(), key=lambda x: x[1], reverse=True)
        status_labels, counts = zip(*sorted_status)
    
        plt.figure(figsize=(5, 3)) # figure size
        plt.bar(status_labels, counts, color="lightblue")
        plt.xlabel('Order status', fontsize=8)
        plt.ylabel('Number of orders', fontsize=8)
        plt.title('Inventory Order Status', fontsize=12)
        
        plt.xticks(fontsize=7)
        plt.yticks(fontsize=7)
   
        figures = plt.gcf()
        mycanvas = FigureCanvasTkAgg(figures, master=charts_frames) #convert it to tkinter canvas
        mycanvas.draw()
        bar_var = mycanvas.get_tk_widget().place(x=20, y=20, width=360)
        plt.close()
                
    def create_pie_chart_sales():
        with open("CREDS/products.json", 'r') as readF:
            read = json.load(readF)
            
        sorted_items = sorted(read["products"], key=lambda sort: sort["Quantity_Sold"], reverse=True) # sort items and reversed to get top 4 itmes
        top_4_sold = sorted_items[:4] # get top 4 itmes
        
        product_names = [prod["Product_Name"] for prod in top_4_sold]
        quantities_sold = [prod["Quantity_Sold"] for prod in top_4_sold]
        
        plt.figure(figsize=(5,3)) # diagram size
        # plt.figure(facecolor="#f2f2f2")
        for i,each in enumerate(top_4_sold,start=0):
            plt.pie(quantities_sold, labels=product_names, autopct='%1.1f%%',startangle=180)
        plt.title('Distribution of Quantities Sold', fontsize=12)
        
        plt.rcParams['font.size'] = 8
        
        figures = plt.gcf()
        pie_canvas = FigureCanvasTkAgg(figures, master=charts_frames)#convert it to tkinter canvas
        pie_canvas.draw()
        pie_canvas.get_tk_widget().place(x=500, y=20, width=350)
        
        plt.close()
        

    # --------------------------------------------------------------------------------------------------------------------------------------------------
    # ORDER PAGE 
    orders_label = Label(orders_page, text="Orders Page", font=("Serif", 32, "bold"),fg="#379777")
    orders_label.place(x=30,y=30) 

    new_order_btn = Button(orders_page, text="New order", font=("Serif", 13), image=new_order, compound=LEFT, command=lambda:show_frames(new_order_page), relief=FLAT)
    new_order_btn.place(x=775, y=30)

    # ------------------------------------------------------------------------------------------------------------------------------------------------------
    orders_frame = ctk.CTkFrame(orders_page, width=250, height=75, fg_color="#379777", border_width=2, border_color="#333333")
    orders_frame.place(x=30, y=120)

    order_icon_1 = ImageTk.PhotoImage(Image.open("ASSETS\\orders_icon_1.png").resize((50, 50)))

    orders_img_lbl = Label(orders_frame, image=order_icon_1, bg="#379777")
    orders_img_lbl.place(x=10, y=10)

    orders_lbl_1 = Label(orders_frame, text="Orders", font=("Serif", 13, "bold"), bg="#379777", fg="#f2f2f2")
    orders_lbl_1.place(x=80,y=10)

    orders_lbl_count = Label(orders_frame, text="0", font=("Serif", 18, "bold"), bg="#379777", fg="#f2f2f2")
    orders_lbl_count.place(x=80,y=30)

    # ------------------------------------------------------------------------------------------------------------------------------------------------------
    shipping_frame = ctk.CTkFrame(orders_page, width=250, height=75, fg_color="#379777", border_width=2, border_color="#333333")
    shipping_frame.place(x=350, y=120)

    order_icon_2 = ImageTk.PhotoImage(Image.open("ASSETS\\shipping_icon_1.png").resize((50, 50)))

    orders_img_lbl_1 = Label(shipping_frame, image=order_icon_2, bg="#379777")
    orders_img_lbl_1.place(x=10, y=10)

    shipping_lbl = Label(shipping_frame, text="Confirmed", font=("Serif", 13, "bold"), bg="#379777", fg="#f2f2f2")
    shipping_lbl.place(x=80,y=10)

    shipping_lbl_count = Label(shipping_frame, text="0", font=("Serif", 18, "bold"), bg="#379777", fg="#f2f2f2")
    shipping_lbl_count.place(x=80,y=30)

    # ------------------------------------------------------------------------------------------------------------------------------------------------------
    delivered_frame = ctk.CTkFrame(orders_page, width=250, height=75, fg_color="#379777", border_width=2, border_color="#333333")
    delivered_frame.place(x=660, y=120)

    order_icon_3 = ImageTk.PhotoImage(Image.open("ASSETS\\delivered_icon_1.png").resize((50, 50)))
    orders_img_lbl = Label(delivered_frame, image=order_icon_3, bg="#379777")
    orders_img_lbl.place(x=10,y=10)

    delivered_lbl = Label(delivered_frame, text="Delivered", font=("Serif", 13, "bold"), bg="#379777", fg="#f2f2f2")
    delivered_lbl.place(x=80,y=10)

    delivered_lbl_count = Label(delivered_frame, text="0", font=("Serif", 18, "bold"), bg="#379777", fg="#f2f2f2")
    delivered_lbl_count.place(x=80,y=30)


    # ------------------------------------------------------------------------------------------------------------------------------------------------------

    search_conatiner = ctk.CTkFrame(orders_page, width=880, height=50, border_width=2, border_color="#333333", fg_color="#f2f2f2")
    search_conatiner.place(x=30,y=220)

    search_entry = ctk.CTkEntry(search_conatiner, font=("Serif", 18), width=770, fg_color="#f2f2f2", text_color="#333333", border_width=2, corner_radius=10, placeholder_text="Search Order", border_color="#379777")
    search_entry.place(x=10, y=10)
 
    search_button = ctk.CTkButton(search_conatiner, text="Search", border_width=2, border_color="#379777", cursor="hand2", width=80, height=30, fg_color="#379777", font=("Serif", 18), hover_color="#36BA98")
    search_button.place(x=795,y=10)


    def search_func():
        '''seach func use to search through the treeview...'''
        with open("CREDS/orders.json", 'r') as readF:
            read = json.load(readF)

        search_1 = search_entry.get().lower()
        
        filtered_data = []

        # filters data the then appends it to the list 
        for row in read["inventory_orders"]:
            # covers all the posible search options using or
            if search_1 in row["Item_Name"].lower() or row["Item_Name"].lower().startswith(search_1) or search_1 in row["Customer"].lower() or search_1 in row["Address"].lower() or search_1 in row["Item_ID"].lower():
                filtered_data.append(row) # append to the list
        
        for item in Dataview.get_children():
            # deletes all the items in the treeveiew
            Dataview.delete(item)
        
        for each in filtered_data:
            # insert the filtered data via for loop
            Dataview.insert("", "end", values=(each["Item_ID"],
                                            each["Item_Name"],
                                            each["Customer"],
                                            each["Address"],
                                            each["Status"],
                                            each["Quantity"]))

    def key_search(event):
        '''
        the first if checks if the entry get nothings if true it will run the UpdateTV()-Update treeviews
        if it get something it will use the search function
        '''
        if search_entry.get() == "":
            update_TV()
        else:
            search_func()

    search_entry.bind("<KeyRelease>", key_search)

    # ---------------------------------------------------------------------------------------------------------------
    # Treeview
    tree_frame = Frame(orders_page)
    tree_frame.place(x=30, y=280)
    tree_frame.pack_propagate(True)

    # styling the treeview
    mycolumns = ("Order ID", "Item Name", "Customer", "Address", "Status", "Quantity")
    Dataview = ttk.Treeview(tree_frame, columns=mycolumns, show="headings", height=15)
    Dataview.pack(side="left", fill="y")

    vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=Dataview.yview) 
    vsb.pack(side="right", fill="y")
    Dataview.configure(yscrollcommand=vsb.set)
    
    # tree view design
    def style_treeview():
        # use to style the treeview
        treeview_style = ttk.Style()
        treeview_style.configure("Treeview.Heading", foreground="#000000", font=("Serif", 10, "bold"), background="red")
        treeview_style.configure("Treeview", background="#f2f2f2", foreground="#000000", rowheight=25, font=("Serif", 10))
        treeview_style.map('Treeview', background=[("selected", "#379777")], foreground=[("selected", "white")])

    
    normal_width = [152, 171, 152, 133, 114, 152] # size of the treeview
    for each in range(len(mycolumns)):
        Dataview.column(mycolumns[each], width=normal_width[each], anchor="center")
        Dataview.heading(mycolumns[each],text=mycolumns[each])
        
    order_menu_order_frame = ctk.CTkFrame(orders_page,border_width=2, border_color="#333333",width=880, height=55, fg_color="#f2f2f2")
    order_menu_order_frame.place(x=30, y=690)
    order_menu_order_frame.grid_propagate(False)
    
    # update_btn_order = ctk.CTkButton(order_menu_order_frame, text="Update", width=190, fg_color="#379777", text_color="#f2f2f2", font=("Serif", 20, "bold"), hover_color="#36BA98", state=DISABLED, command=lambda: update_order_func())
    # update_btn_order.grid(row=0, column=0, padx=13, pady=10)

    delete_btn_order = ctk.CTkButton(order_menu_order_frame, text="Delete", width=260, fg_color="#379777", text_color="#f2f2f2", font=("Serif", 20, "bold"), hover_color="#36BA98", command=lambda:delete_order())
    delete_btn_order.grid(row=0, column=0, padx=15, pady=10)

    edit_btn_order = ctk.CTkButton(order_menu_order_frame, text="Edit", width=260, fg_color="#379777", text_color="#f2f2f2", font=("Serif", 20, "bold"), hover_color="#36BA98", command=lambda:edit_order_func())
    edit_btn_order.grid(row=0, column=1, padx=15, pady=10)

    export_btn_order = ctk.CTkButton(order_menu_order_frame, text="Export to CSV/Excel", width=260, fg_color="#379777", text_color="#f2f2f2", font=("Serif", 20, "bold"), hover_color="#36BA98", command=lambda:export_to_what_inventory())
    export_btn_order.grid(row=0, column=2, padx=15, pady=10)

    # order page functions
    def delete_order():
        '''Ask yer or no if yes it will delete the clicked row else it will return None'''
        try:
            if messagebox.askyesno("ALERT","Do you want to delete this row?"):
                with open("CREDS/orders.json", 'r') as readF:
                    data = json.load(readF)

                click = Dataview.focus()
                clicked = Dataview.item(click)["values"]
                Dataview.delete(click)

                filtered_data = []
                for each in data["inventory_orders"]:
                    if clicked[1] != each["Item_Name"]:
                        filtered_data.append(each)

                new_data = {"inventory_orders":filtered_data}
                with open("CREDS/orders.json", 'w') as readW:
                    json.dump(new_data, readW, indent=4)
            else:
                return None
        except:
            messagebox.showerror("ALERT!", "You haven't select anything.")

    def edit_order_func():
        global item_ID_ent_1, item_name_ent_1, item_customer_ent, item_address_ent, value_var, edit_order_win,quantity_box_1
        '''Makes a new page and insert the selected table in the treview'''
        try:
            edit_order_win = Toplevel()
            edit_order_win.geometry("500x730")
            edit_order_win.title("Edit")
            edit_order_win.resizable(False, False)

            def return_normal(event):
                edit_btn_order.configure(state="normal")

           
            edit_btn_order.configure(state="disabled")

            def insert_Datas_ordersPG():
                click = Dataview.focus()
                clicked = Dataview.item(click)["values"]
                item_ID_ent_1.insert(0, clicked[0])
                item_name_ent_1.insert(0, clicked[1])
                item_customer_ent.insert(0, clicked[2])
                item_address_ent.insert(0, clicked[3])
                value_var.set(clicked[4])
                quantity_box_1.insert(0, clicked[5])
     

            Label(edit_order_win, text="Edit Order", font=("Serif", 20, "bold")).place(x=30, y=30)

            Label(edit_order_win, text="Item ID", fg="#379777", font=("Serif",14)).place(x=30, y=100)

            item_ID_ent_1 = ctk.CTkEntry(edit_order_win, font=("Serif", 18, "bold"), fg_color="#E0E0E0",border_color="#E0E0E0", corner_radius=5, height=40, width=450, text_color="#000000")
            item_ID_ent_1.place(x=30, y=130)

            Label(edit_order_win, text="Item Name", fg="#379777", font=("Serif",14)).place(x=30, y=180)

            item_name_ent_1 = ctk.CTkEntry(edit_order_win, font=("Serif", 18, "bold"), fg_color="#E0E0E0",border_color="#E0E0E0", corner_radius=5, height=40, width=450, text_color="#000000")
            item_name_ent_1.place(x=30, y=210)

            Label(edit_order_win, text="Customer", fg="#379777", font=("Serif",14)).place(x=30, y=260)

            item_customer_ent = ctk.CTkEntry(edit_order_win, font=("Serif", 18, "bold"), fg_color="#E0E0E0",border_color="#E0E0E0", corner_radius=5, height=40, width=450, text_color="#000000")
            item_customer_ent.place(x=30, y=290)

            Label(edit_order_win, text="Address", fg="#379777", font=("Serif",14)).place(x=30, y=340)

            item_address_ent = ctk.CTkEntry(edit_order_win, font=("Serif", 18, "bold"), fg_color="#E0E0E0",border_color="#E0E0E0", corner_radius=5, height=40, width=450, text_color="#000000")
            item_address_ent.place(x=30, y=370)

            Label(edit_order_win, text="Status", fg="#379777", font=("Serif",14)).place(x=30, y=420)

            value_var = StringVar()

            confirm_rbtn = ctk.CTkRadioButton(edit_order_win, text="Confirmed", fg_color="#379777", font=("Serif", 18, "bold"), value="Confirmed", variable=value_var, text_color="#379777", border_color="#379777")
            confirm_rbtn.place(x=70, y=460)

            pending_rbtn = ctk.CTkRadioButton(edit_order_win, text="Delivered", fg_color="#379777", font=("Serif", 18, "bold"), value="Delivered", variable=value_var, text_color="#379777", border_color="#379777")
            pending_rbtn.place(x=70, y=500)

            cancelled_rbtn = ctk.CTkRadioButton(edit_order_win, text="Cancelled", fg_color="#379777", font=("Serif", 18, "bold"), value="Cancelled", variable=value_var, text_color="#379777", border_color="#379777")
            cancelled_rbtn.place(x=70, y=540)

            Label(edit_order_win, text="Quantity", fg="#379777", font=("Serif",14)).place(x=30, y=580)

            quantity_box_1 = ctk.CTkEntry(edit_order_win, font=("Serif", 18, "bold"), fg_color="#E0E0E0",border_color="#E0E0E0", corner_radius=5, height=40, width=450, text_color="#000000")
            quantity_box_1.place(x=30, y=620)

            update_btn_order = ctk.CTkButton(edit_order_win, text="Update", fg_color="#379777", font=("Serif", 18, "bold"), command=lambda: update_order_func())
            update_btn_order.place(x=180,y=680)

            insert_Datas_ordersPG()

            edit_order_win.bind("<Destroy>", return_normal)

        except IndexError:
            messagebox.showerror("ERROR", "You haven't selected anything yet.")
            edit_order_win.bind("<Destroy>", return_normal)
            edit_order_win.destroy()
            # update_btn_order.configure(state="disabled")
            edit_btn_order.configure(state="normal")
            return


        edit_order_win.mainloop()



    def update_order_func():
        '''update function in the order that iterates over the invetory orders then change the value with what the user input.'''
        if messagebox.askyesno("UPDATE", "Do you want to update this?:"):
            with open("CREDS/orders.json", 'r') as readF:
                read = json.load(readF)

            click1 = Dataview.selection()
            clicked1 = Dataview.item(click1)["values"]

            for each in read["inventory_orders"]:
                if each["Item_Name"] == clicked1[1]:
                        each["Item_ID"] = item_ID_ent_1.get()
                        each["Item_Name"] = item_name_ent_1.get()
                        each["Customer"] = item_customer_ent.get()
                        each["Address"] = item_address_ent.get()
                        each["Status"] = value_var.get()
                        each["Quantity"] = quantity_box_1.get()


            with open("CREDS/orders.json", 'w') as writeF:
                json.dump(read, writeF, indent=4)
            
            edit_btn_order.configure(state="normal")
            update_TV()
            filter_data()
            edit_order_win.destroy()

        else:
            return None            
    
    def export_to_what_order():
        '''Ask user yes and no to generate xlsx report and PDF report..... Yes for xlsx and No for PDF report'''
        dialouge = messagebox.askyesnocancel("Export to What?","Select Yes to Export Via CSV/Excel, No for Via PDF report, and Cancel to quit")
        if dialouge:
            '''Create a excel report sheet'''
            with open("CREDS/orders.json", 'r') as file:
                read = json.load(file)
            wB = Workbook()
            sheet = wB.active
            sheet.title = "Sales Report"
            file_name = simpledialog.askstring("Inuput", "Please enter your chosen file name: ")
            
            headers = [
                "Order ID","Item Name",
                "Customer", "Address",
                "Status", "Quantity"
            ]
            sheet.append(headers)
            
            colum_widt = [len(headers[0]), len(headers[1]), len(headers[2]), len(headers[3]),len(headers[4]),len(headers[5])]
            
            for i, width in enumerate(colum_widt, start=1):
                column_lt = get_column_letter(i)
                sheet.column_dimensions[column_lt].width = width * 3

            center_allign = Alignment(horizontal="center", vertical="center") 
            for cell in sheet[1]:
                cell.alignment = center_allign
                               
            for each in read["inventory_orders"]:
                sheet.append([
                    each["Item_ID"], each["Item_Name"],
                    each["Customer"], each["Address"],
                    each["Status"], each["Quantity"]
                ])
    
            path_name = path_lbl.cget("text")
            wB.save(f"{path_name}\\{file_name}.xlsx")
            messagebox.showinfo("ALERT",f"file saved at: {path_name}\\{file_name}.xlsx")
                
        elif dialouge == False:
            '''generate pdf report'''
            with open("CREDS/orders.json", 'r') as file:
                read = json.load(file)
                
            file_name = simpledialog.askstring("Inuput", "Please enter your chosen file name: ")
            path_name  = path_lbl.cget("text")
            pdf_file_name = f"{path_name}\\{file_name}.pdf"
            doc = SimpleDocTemplate(pdf_file_name, pagesize=landscape(letter))
            
            elements = []
            styles = getSampleStyleSheet()
            
            
            centered_text = Paragraph("Sales Report", styles['Title'])
            elements.append(centered_text)
            elements.append(Spacer(1, 12))  # Add a spacer for spacing
            
            
            table_data = []
            headers = ["Order ID", "Item Name", "Customer", "Address", "Status", "Quantity"]
            table_data.append(headers)
            
            for each in read["inventory_orders"]:
                pdf_rows = [
                    each["Item_ID"],
                    each["Item_Name"],
                    each["Customer"],
                    each["Address"],
                    each["Status"],
                    each["Quantity"]
                ]
                table_data.append(pdf_rows)
            
            pdf_table = Table(table_data)
            
            style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ])
            pdf_table.setStyle(style)
            elements.append(pdf_table)
            doc.build(elements)
            messagebox.showinfo("ALERT!", f"File saved at {path_name}\\{file_name}.pdf")
        else:
            return None    
    # end
    
    # CREATE ORDER
    make_order_lbl = Label(new_order_page, text="Create order", font=("Serif", 32, "bold"), bg="#f2f2f2", fg="#379777")
    make_order_lbl.place(x=30, y=30)

    item_ID_lbl = Label(new_order_page, text="Item ID", font=("Serif", 16),bg="#f2f2f2", fg="#379777")
    item_ID_lbl.place(x=30, y=130)

    item_ID_ent = ctk.CTkEntry(new_order_page, font=("Serif", 16), width=880, fg_color="#E0E0E0", text_color="#000000", border_color="#E0E0E0", height=35)
    item_ID_ent.place(x=30, y=160)

    item_name_lbl = Label(new_order_page, text="Item Name", font=("Serif",16),bg="#f2f2f2", fg="#379777")
    item_name_lbl.place(x=30, y=200)

    item_name_ent = ctk.CTkEntry(new_order_page, font=("Serif", 16), width=880, fg_color="#E0E0E0", text_color="#000000", border_color="#E0E0E0", height=35)
    item_name_ent.place(x=30, y=230)

    customer_lbl = Label(new_order_page, text="Customer",font=("Serif", 14),bg="#f2f2f2", fg="#379777")
    customer_lbl.place(x=30, y=270)

    customer_ent = ctk.CTkEntry(new_order_page, font=("Serif", 16), width=430, fg_color="#E0E0E0", text_color="#000000", border_color="#E0E0E0", height=35)
    customer_ent.place(x=30, y=300)

    Address_lbl = Label(new_order_page, text="Address",font=("Serif", 14),bg="#f2f2f2", fg="#379777")
    Address_lbl.place(x=485, y=270)

    address_ent = ctk.CTkEntry(new_order_page, font=("Serif", 16), width=430, fg_color="#E0E0E0", text_color="#000000", border_color="#E0E0E0", height=35)
    address_ent.place(x=485, y=300)

    status_lbl = Label(new_order_page, font=("Serif", 14), text="Status",bg="#f2f2f2", fg="#379777")
    status_lbl.place(x=30, y=350)

    status_var = StringVar()
    confirm_btn = ctk.CTkRadioButton(new_order_page, text="Confirmed", fg_color="#379777", font=("Serif", 18), value="Confirmed", variable=status_var, text_color="#379777", border_color="#379777")
    confirm_btn.place(x=30, y=390)

    delivered_btn = ctk.CTkRadioButton(new_order_page, text="Delivered", fg_color="#379777", font=("Serif", 18), value="Delivered", variable=status_var, text_color="#379777", border_color="#379777")
    delivered_btn.place(x=30, y=440)

    cancelled_btn = ctk.CTkRadioButton(new_order_page, text="Cancelled", fg_color="#379777", font=("Serif", 18), value="Cancelled", variable=status_var, text_color="#379777", border_color="#379777")
    cancelled_btn.place(x=30, y=490)

    Description_lbl = Label(new_order_page, text="Description", font=("Serif", 14),bg="#f2f2f2", fg="#379777")
    Description_lbl.place(x=485, y=350)

    Description_box = ctk.CTkTextbox(new_order_page, font=("Serif", 18), width=430, height=225, fg_color="#E0E0E0", text_color="#000000", border_color="#E0E0E0")
    Description_box.place(x=485, y=390)

    quantity_lbl = Label(new_order_page, text="Quantity", font=("Serif", 14),bg="#f2f2f2", fg="#379777")
    quantity_lbl.place(x=30, y=540)

    quantity_box = ctk.CTkEntry(new_order_page, font=("Serif", 14),width=430, fg_color="#E0E0E0", text_color="#000000", border_color="#E0E0E0", height=35)
    quantity_box.place(x=30,y=580)

    create_btn = ctk.CTkButton(new_order_page, text="Create Order", font=("Serif", 20,"bold"),width=350,height=40, cursor="hand2", fg_color="#379777", command=lambda:create_order())
    create_btn.place(x=270, y=670)
    # end

    # ---INVENTORY---------------------------------------------------------------------------------------------------------------------------------

    inventory_label = Label(inventory_page, text="Inventory Page", font=("Serif", 32, "bold"),fg="#379777")
    inventory_label.place(x=30,y=20) 

    product_frame = ctk.CTkFrame(inventory_page, fg_color="#379777", width=400, height=125)
    product_frame.place(x=30, y=80)

    out_of_stocks = Label(product_frame, text="Product Details: ", font=("Serif", 20, "bold"), bg="#379777", fg="#f2f2f2").place(x=20, y=10)

    out_of_stocks = Label(product_frame, text="Out of stocks Items: ", font=("Serif",15, "bold"), fg="#f2f2f2", bg="#379777").place(x=20, y=50)

    out_of_stocks_count = Label(product_frame, text="0", font=("Serif",15, "bold"), fg="#FF0000", bg="#379777")
    out_of_stocks_count.place(x=230, y=50)

    total_items = Label(product_frame, text="Total Items: ", font=("Serif",15, "bold"), bg="#379777", fg="#f2f2f2").place(x=20, y=80)

    total_items_count = Label(product_frame, text="0", font=("Serif",15, "bold"), bg="#379777", fg="#f2f2f2")
    total_items_count.place(x=230, y=80)

    top_selling_frame = ctk.CTkFrame(inventory_page, fg_color="#379777", width=400, height=125)
    top_selling_frame.place(x=500, y=80)

    top_selling_lbl = Label(top_selling_frame, text="Top selling products: ", font=("Serif", 13, "bold"), bg="#379777", fg="#f2f2f2")
    top_selling_lbl.place(x=10, y=0)

    uno = Label(top_selling_frame, text="1:", font=("Serif", 10, "bold"), bg="#379777", fg="#f2f2f2")
    uno.place(x=10, y=25)

    dos = Label(top_selling_frame, text="2:", font=("Serif", 10, "bold"), bg="#379777", fg="#f2f2f2")
    dos.place(x=10, y=50)

    tres = Label(top_selling_frame, text="3:", font=("Serif", 10, "bold"), bg="#379777", fg="#f2f2f2")
    tres.place(x=10, y=75)

    quatro = Label(top_selling_frame, text="4:", font=("Serif", 10, "bold"), bg="#379777", fg="#f2f2f2")
    quatro.place(x=10, y=100)

    #
    
    search_conatiner_1 = ctk.CTkFrame(inventory_page, width=880, height=50, border_width=2, border_color="#333333", fg_color="#f2f2f2")
    search_conatiner_1.place(x=30,y=220)

    search_entry_1 = ctk.CTkEntry(search_conatiner_1, font=("Serif", 18), width=770, fg_color="#f2f2f2", text_color="#333333", border_width=2, corner_radius=10, placeholder_text="Search Order", border_color="#379777")
    search_entry_1.place(x=10, y=10)

    search_button_1 = ctk.CTkButton(search_conatiner_1, text="Search", border_width=2, border_color="#379777", cursor="hand2", width=80, height=30, fg_color="#379777", font=("Serif", 18), hover_color="#36BA98")
    search_button_1.place(x=795,y=10)

    #
    
    tree_frame = Frame(inventory_page)
    tree_frame.place(x=30, y=280)

    mycolumns = ("Product Name", "Quantity", "Price", "Number of Sold", "Cost")
    inventory_tv = ttk.Treeview(tree_frame, columns=mycolumns, show="headings", height=15)
    inventory_tv.pack(side="left", fill="y")

    vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=inventory_tv.yview)
    vsb.pack(side="right", fill="y")

    inventory_tv.configure(yscrollcommand=vsb.set)

    for each in mycolumns:
        inventory_tv.column(each, width=(len(each) *20), anchor="center")
        inventory_tv.heading(each,text=each)

    bottom_menu_inventory_frame = ctk.CTkFrame(inventory_page,border_width=2, border_color="#333333",width=880, height=55, fg_color="#f2f2f2")
    bottom_menu_inventory_frame.place(x=30, y=690)
    bottom_menu_inventory_frame.grid_propagate(False)

    update_btn_inventory = ctk.CTkButton(bottom_menu_inventory_frame, text="Update", width=190, fg_color="#379777", text_color="#f2f2f2", font=("Serif", 20, "bold"), hover_color="#36BA98", state=DISABLED, command=lambda:update_invetory_btn())
    update_btn_inventory.grid(row=0, column=0, padx=15, pady=10)

    delete_btn_inventory = ctk.CTkButton(bottom_menu_inventory_frame, text="Delete",width=190, fg_color="#379777", text_color="#f2f2f2", font=("Serif", 20, "bold"), hover_color="#36BA98", command=lambda: delete_inventory_btn())
    delete_btn_inventory.grid(row=0, column=1, padx=10, pady=10)

    edit_btn_inventory_1 = ctk.CTkButton(bottom_menu_inventory_frame, text="Edit",width=190, fg_color="#379777", text_color="#f2f2f2", font=("Serif", 20, "bold"), hover_color="#36BA98", command=lambda:edit_window_inventory())
    edit_btn_inventory_1.grid(row=0, column=2, padx=10, pady=10)

    export_btn_inventory = ctk.CTkButton(bottom_menu_inventory_frame, text="Export to CSV/Excel",width=190, fg_color="#379777", text_color="#f2f2f2", font=("Serif", 20, "bold"), hover_color="#36BA98", command=lambda:export_to_what_order())
    export_btn_inventory.grid(row=0, column=3, padx=10, pady=10)

    def edit_window_inventory():
        global product_name_entry_edit, quantity_entry_edit, price_entry_edit,number_ofSold_entry_edit, edit_window
        try:
            update_btn_inventory.configure(state="normal")
            edit_btn_inventory_1.configure(state="disabled")

            edit_window = Toplevel()
            edit_window.title("Edit Window")
            edit_window.geometry("500x350")
            edit_window.config(bg="#f2f2f2")

            Frame(edit_window, bg="#59CE8F", width=800,height=30).place(x=0,y=0)

            def return_normal(event):
                edit_btn_inventory_1.configure(state="normal")
                update_btn_inventory.configure(state="disabled")


            def reset_entries():
                product_name_entry_edit.delete(0, END)
                quantity_entry_edit.delete(0, END)
                price_entry_edit.delete(0, END)
                number_ofSold_entry_edit.delete(0, END)


            product_name_lbl_edit = Label(edit_window, text="Product name:", font=("Serif",13,"bold"))
            product_name_lbl_edit.place(x=20, y=30)

            product_name_entry_edit = ttk.Entry(edit_window, font=("Serif", 13,"bold"), width=50)
            product_name_entry_edit.place(x=20, y=60)

            quantity_lbl_edit = Label(edit_window, text="Quantity:", font=("Serif",13,"bold"))
            quantity_lbl_edit.place(x=20, y=90)

            quantity_entry_edit = ttk.Entry(edit_window, font=("Serif", 13,"bold"), width=50)
            quantity_entry_edit.place(x=20, y=120)

            price_lbl_edit = Label(edit_window, text="Price", font=("Serif",13,"bold"))
            price_lbl_edit.place(x=20, y=150)

            price_entry_edit = ttk.Entry(edit_window, font=("Serif", 13,"bold"), width=50)
            price_entry_edit.place(x=20, y=180)

            number_ofSold_lbl_edit = Label(edit_window, text="Number of Sold", font=("Serif",13,"bold"))
            number_ofSold_lbl_edit.place(x=20, y=210)

            number_ofSold_entry_edit = ttk.Entry(edit_window, font=("Serif", 13,"bold"), width=50)
            number_ofSold_entry_edit.place(x=20, y=240)

            edit_reset = Button(edit_window, text="Reset", font=("Serif", 13, "bold"), bg="#59CE8F", fg="#f2f2f2", command=lambda:reset_entries())
            edit_reset.place(x=400, y=280)

            reset_entries()

            click = inventory_tv.focus()
            clicked = inventory_tv.item(click)["values"]

            product_name_entry_edit.insert(0,[clicked[0]])
            quantity_entry_edit.insert(0, clicked[1])
            price_entry_edit.insert(0, clicked[2])
            number_ofSold_entry_edit.insert(0, clicked[3])
        except IndexError:
            messagebox.showerror("ERROR", "You haven't selected anything yet.")
            edit_window.destroy()
            edit_btn_inventory_1.configure(state="normal")
            update_btn_inventory.configure(state="disabled")
            return

        edit_window.bind("<Destroy>", return_normal)

        edit_window.mainloop()

    def update_invetory_btn():
        if messagebox.askyesno("UPDATE", "Do you want to update this?:"):
            with open("CREDS/products.json", 'r') as readF:
                read = json.load(readF)

            click1 = inventory_tv.selection()
            clicked1 = inventory_tv.item(click1)["values"]


            for each in read["products"]:
                if each["Product_Name"] == clicked1[0]:
                    each["Product_Name"] = product_name_entry_edit.get()
                    each["Quantity"] = int(quantity_entry_edit.get())
                    each["Price"] = float(price_entry_edit.get())
                    each["Quantity_Sold"] = int(number_ofSold_entry_edit.get())

            with open("CREDS/products.json", 'w') as writeF:
                json.dump(read, writeF, indent=4)
            
            update_btn_inventory.configure(state="disabled")
            edit_btn_inventory_1.configure(state="normal")
            update_TV()
            filter_data()
            create_pie_chart_sales()
            create_order_status()
            edit_window.destroy()
        else:
            return None
        
    def delete_inventory_btn():
        try:
            if messagebox.askyesno("ALERT","Do you want to delete this row?"):
                with open("CREDS/products.json", "r") as readF:
                    data = json.load(readF)

                click = inventory_tv.focus()
                clicked = inventory_tv.item(click)["values"]
                inventory_tv.delete(click)


                filtered_data = []
                for each in data["products"]:
                    if clicked[0] != each["Product_Name"]:
                        filtered_data.append(each)

                new_data = {"products":filtered_data}
                with open("CREDS/products.json", "w") as readW:
                    json.dump(new_data, readW, indent=4)
            else:
                return None
        except:
            messagebox.showerror("ALERT!", "You haven't select aynthing.")

    def inventory_search():
        with open("CREDS/products.json", 'r') as readF:
            read = json.load(readF)

        search_inventory = search_entry_1.get().lower()
        
        filtered_data = []
        for row in read["products"]:
            if search_inventory  in row["Product_Name"].lower() or row["Product_Name"].lower().startswith(search_inventory):
                filtered_data.append(row)
        
        for item in inventory_tv.get_children():
            inventory_tv.delete(item)
        
        for each in filtered_data:
            inventory_tv.insert("", "end", values=(each["Product_Name"],
                                                each["Quantity"],
                                                each["Price"],
                                                each["Quantity_Sold"]))

    def key_search_inventory(event):
        if search_entry_1.get() == "":
            update_TV()
        else:
            inventory_search()

    search_entry_1.bind("<KeyRelease>", key_search_inventory)
    
    def clear_order_wigets():
        item_ID_ent.delete(0, END)
        item_name_ent.delete(0, END)
        customer_ent.delete(0, END)
        address_ent.delete(0, END)
        Description_box.delete(1.0, END)
        status_var.set("")
        quantity_box.delete(0, END)

    def update_TV():
        with open ("CREDS\\orders.json", 'r') as readF:
            read = json.load(readF)
        
        for each in Dataview.get_children():
            Dataview.delete(each)
        
        for each in read["inventory_orders"]:
            Dataview.insert("", "end", values=(each["Item_ID"], each["Item_Name"], each["Customer"], each["Address"],each["Status"], each["Quantity"]))
        # ----------------------------------------
        with open ("CREDS\\products.json", 'r') as readF:
            read1 = json.load(readF)
            
        for each1 in inventory_tv.get_children():
            inventory_tv.delete(each1)
            
        for each1 in read1["products"]:
            inventory_tv.insert("","end",values=(each1["Product_Name"], each1["Quantity"], each1["Price"], each1["Quantity_Sold"], each1["Cost"]))
            
        create_order_status()
        create_pie_chart_sales()
        
    def filter_data():
        """
        Fetches and filters data from JSON files to display various metrics and calculations.

        Retrieves order status data from 'orders.json' and product inventory data from 'products.json'.
        Calculates total number of orders, items out of stock, total sales revenue, total cost, and total earnings.
        Displays these metrics in GUI elements and prints total cost, total sales, and total earnings to console.
        """
        ORDERS = 0
        SHIPPING = 0
        DELIVERED = 0
        PENDING = 0
        CANCELLED = 0
        CONFIRMED = 0
        
        with open("CREDS\\orders.json", 'r') as readF:
            read = json.load(readF)
        
        for each in read["inventory_orders"]:
            ORDERS += 1
            if each["Status"] == "Shipped":
                SHIPPING += 1
            elif each["Status"] == "Pending":
                PENDING += 1 
            elif each["Status"] == "Delivered":
                DELIVERED += 1
            elif each["Status"] == "Cancelled":
                CANCELLED += 1
            elif each["Status"] == "Confirmed":
                CONFIRMED += 1
                
        orders_lbl_count.config(text=ORDERS)
        shipping_lbl_count.config(text=CONFIRMED)
        delivered_lbl_count.config(text=DELIVERED)
        
        TOTAL_ITEM = 0
        OUT_STOCKS = 0
        TOTAL_SALES = 0
        TOTAL_COST = 0

        with open("CREDS\\products.json", 'r') as readF:
            read1 = json.load(readF)
        
        for each in read1["products"]:
            TOTAL_ITEM += 1

            revenue = each["Quantity_Sold"] * each["Price"]
            TOTAL_SALES += revenue

            cost = each["Quantity_Sold"] * each["Cost"]
            TOTAL_COST += cost

            if each["Quantity"] < each["Quantity_Sold"]:
                OUT_STOCKS += 1
                
        total_items_count.config(text=TOTAL_ITEM)
        out_of_stocks_count.config(text=OUT_STOCKS)

        # GET TOP 4 MOST SOLD
        sorted_items = sorted(read1["products"], key=lambda sort: sort["Quantity_Sold"], reverse=True)
        top_4_sold = sorted_items[:4]
        
        uno.config(text=f"{top_4_sold[0]['Product_Name']}: {top_4_sold[0]['Quantity_Sold']}")
        dos.config(text=f"{top_4_sold[1]['Product_Name']}: {top_4_sold[1]['Quantity_Sold']}")
        tres.config(text=f"{top_4_sold[2]['Product_Name']}: {top_4_sold[2]['Quantity_Sold']}")
        quatro.config(text=f"{top_4_sold[3]['Product_Name']}: {top_4_sold[3]['Quantity_Sold']}")

        TOTAL_EARNINGS = TOTAL_SALES - TOTAL_COST

        total_item_label_counter.config(text=TOTAL_ITEM)
        total_sales_counter_lbl.config(text=TOTAL_SALES)
        revenue_lbl_counter.config(text=TOTAL_EARNINGS)
  
    def create_order():
        data = {
            "Item_ID": item_ID_ent.get(),
            "Item_Name": item_name_ent.get(),
            "Customer": customer_ent.get(),
            "Address": address_ent.get(),
            "Status": status_var.get(),
            "Quantity": quantity_box.get()
        }
        
        try:
            with open("CREDS\\orders.json", 'r') as readF:
                try:
                    read = json.load(readF)
                except json.JSONDecodeError:
                    read = {"inventory_orders": []}
        except FileNotFoundError:
            read = {"inventory_orders": []}

        read["inventory_orders"].append(data)
        
        with open("CREDS\\orders.json", 'w') as readW:
            json.dump(read, readW, indent=4)
        clear_order_wigets()
        messagebox.showinfo("Alert!", "Successfully Created!")
        filter_data()
        update_TV()

    def export_to_what_inventory():
        dialouge = messagebox.askyesnocancel("Export to What?","Select Yes to Export Via CSV/Excel, No for Via PDF report, and Cancel to quit")
        if dialouge:
            with open("CREDS/products.json", 'r') as file:
                read = json.load(file)
            wB = Workbook()
            sheet = wB.active
            sheet.title = "Sales Report"
            
            file_name = simpledialog.askstring("Inuput", "Please enter your chosen file name: ")


            headers = [
                "Product_Name","Quantity",
                "Price", "Quantity_Sold",
                "Cost"
            ]
            sheet.append(headers)
            
            colum_widt = [len(headers[0]), len(headers[1]), len(headers[2]), len(headers[3]),len(headers[4])]
            
            for i, width in enumerate(colum_widt, start=1):
                column_lt = get_column_letter(i)
                sheet.column_dimensions[column_lt].width = width * 3

            center_allign = Alignment(horizontal="center", vertical="center") 
            for cell in sheet[1]:
                cell.alignment = center_allign
            
            for each in read["products"]:
                sheet.append([
                    each["Product_Name"], each["Quantity"],
                    each["Price"], each["Quantity_Sold"],
                    each["Cost"]
                ])
                
            path_name  = path_lbl.cget("text")    
            wB.save(f"{path_name}\\{file_name}.xlsx")
            messagebox.showinfo("ALERT!", f"File saved at {path_name}\\{file_name}.xlsx")
                
        elif dialouge == False:
            with open("CREDS/products.json", 'r') as file:
                read = json.load(file)
            file_name = simpledialog.askstring("Inuput", "Please enter your chosen file name: ")

            path_name  = path_lbl.cget("text")
            pdf_file_name = f"{path_name}\\{file_name}.pdf"
            doc = SimpleDocTemplate(pdf_file_name, pagesize=landscape(letter))
            
            elements = []
            styles = getSampleStyleSheet()
            
            centered_text = Paragraph("Sales Report", styles['Title'])
            elements.append(centered_text)
            elements.append(Spacer(1, 12))  # Add a spacer for spacing
            
            table_data = []
            headers = ["Product Name", "Quantity", "Price", "Number Sold", "Cost"]
            table_data.append(headers)
            
            for each in read["products"]:
                pdf_row = [
                    each["Product_Name"], each["Quantity"],
                    each["Price"], each["Quantity_Sold"],
                    each["Cost"]
                ]
                table_data.append(pdf_row)
            
            pdf_table = Table(table_data)
            
            style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ])
            pdf_table.setStyle(style)
            elements.append(pdf_table)
            doc.build(elements)
            messagebox.showinfo("ALERT!", f"File saved at {path_name}\\{file_name}.xlsx")
            
        else:
            return None

        
    def exit():
        if messagebox.askyesno("ALERT!", "Do you want to exit?"):
            root.destroy()
            tst.deiconify()
        else:
            return None


    create_order_status()
    create_pie_chart_sales()
    show_frames(dashboard_page)
    root.bind("<F5>", updater)
    universal_path()
    root.protocol("WM_DELETE_WINDOW", exit)
    update_TV()
    filter_data()
    style_treeview()
    root.mainloop()
        
def register_window(root):
    root.withdraw()
    new_win = Toplevel()
    new_win.title("Register Page")

    app_Width_1, app_Height_1 = 500, 530
    new_win.resizable(False,False)

    screen_Width_1, screen_Height_1 = new_win.winfo_screenwidth(), new_win.winfo_screenheight()
    x, y = (screen_Width_1 / 2) - (app_Width_1 / 2), (screen_Height_1 / 2) - (app_Height_1 / 2)
    new_win.geometry(f"{app_Width_1}x{app_Height_1}+{int(x)}+{int(y)}")

    open_eye_img = Image.open("ASSETS\\eye_open.png")
    close_eye_img = Image.open("ASSETS\\eye_close.png")

    new_win.open_eye = ctk.CTkImage(light_image=open_eye_img, size=(25,25))
    new_win.close_eye = ctk.CTkImage(light_image=close_eye_img, size=(25,25))

    is_on = True
    def show_password():
        nonlocal is_on
        if is_on:
            password_entry_1.configure(show="")
            password_entry_2.configure(show="")
            show_password_btn.configure(image=new_win.open_eye)
            is_on = False
        else:
            password_entry_1.configure(show="•")
            password_entry_2.configure(show="•")
            show_password_btn.configure(image=new_win.close_eye)
            is_on = True

    Label(new_win, text="Create an Account", font=("Serif", 23,"bold"),fg="#333333").place(x=30,y=30)

    name_label = Label(new_win, text="Username", font=("Serif", 16),fg="#333333").place(x=30, y=100)
    user_name = ctk.CTkEntry(new_win, font=("Serif", 16), width=445, height=40,fg_color="#379777", border_color="#379777",text_color="#F4F3F2", placeholder_text="Username", placeholder_text_color="#A9A9AC")
    user_name.place(x=30, y=130)

    password_label_1 = Label(new_win, text="Passsword", font=("Serif", 16),fg="#333333").place(x=30, y=180)
    password_entry_1 = ctk.CTkEntry(new_win, font=("Serif", 16), width=445, height=40,fg_color="#379777", border_color="#379777",text_color="#F4F3F2",placeholder_text="Password", placeholder_text_color="#A9A9AC", show="•")
    password_entry_1.place(x=30, y=210)

    password_label_2 = Label(new_win, text="Confirm Password", font=("Serif", 16),fg="#333333").place(x=30, y=260)
    password_entry_2 = ctk.CTkEntry(new_win, font=("Serif", 16), width=445, height=40,fg_color="#379777", border_color="#379777",text_color="#F4F3F2",placeholder_text="Confirm password", placeholder_text_color="#A9A9AC", show="•")
    password_entry_2.place(x=30, y=290)

    def test_cb():
        if chVariable.get() == "Yes":
            create_account_btn.configure(state="normal")
        else:
            create_account_btn.configure(state="disabled")

    show_password_btn = ctk.CTkButton(new_win, text="Show password", font=("Serif", 16, "bold"),image=new_win.close_eye, command=lambda:show_password())
    show_password_btn.place(x=300, y=350)

    profile_pic = ctk.CTkButton(new_win, text="Upload Profile", font=("Serif", 16, "bold"), height=40, command=lambda:upload_pfp_path())
    profile_pic.place(x=30, y=400)

    profile_pic_path = ctk.CTkLabel(new_win, text="~", font=("Serif", 14, "bold"), height=40, fg_color="#379777", width=300, corner_radius=10)
    profile_pic_path.place(x=180, y=400)

    def upload_pfp_path():
        home_dir = os.path.expanduser('~')
        init_dir = os.path.join(home_dir,"Pictures")
        file_name = filedialog.askopenfilename(initialdir=init_dir)
        if file_name:
            profile_pic_path.configure(text=file_name)
        else:
            return None

    admin_varch = StringVar()

    admin_checker = ctk.CTkCheckBox(new_win, text="Are you admin?", fg_color="#379777", text_color="#333333", variable=admin_varch, onvalue="Yes", offvalue="No", command=lambda:test_cb())
    admin_checker.place(x=350, y=450)


    chVariable = StringVar()

    validator_cb = ctk.CTkCheckBox(new_win, text="I remember my password", fg_color="#379777", text_color="#333333", variable=chVariable, onvalue="Yes", offvalue="No", command=lambda:test_cb())
    validator_cb.place(x=30, y=350)

    create_account_btn = ctk.CTkButton(new_win, text="Create Account", font=("Serif", 16, "bold"), width=445, height=40, state=DISABLED, command=lambda:registered())
    create_account_btn.place(x=30, y=480)


    def clear_widgets():
        user_name.delete(0, END)
        password_entry_1.delete(0, END)
        password_entry_2.delete(0, END)

    def registered():
        if admin_varch.get() == "Yes":
            question = simpledialog.askstring("Secret", "Who is the creator of this software?")
            
            if question == "Jerwin" or question == "Oliver":
                fetch_data = {
                    "User_name": user_name.get(),
                    "Password": encrypt_pw(password_entry_1.get()),
                    "User_type": "Admin",
                    "Profile_path": profile_pic_path.cget("text"),
                    "is_logged_in": None
                }

                if password_entry_1.get() != password_entry_2.get():
                    messagebox.showwarning("Read", "Your password doesn't match.")
                    return
                    
                if os.path.exists("CREDS/user_creds.json"):
                    with open("CREDS/user_creds.json", 'r') as rfile:
                        data = json.load(rfile)
                else:
                    data = {"user_credentials": []}
                    with open("CREDS/user_creds.json", 'w') as wfile:
                        json.dump(data, wfile, indent=4)

                validator = [samp for samp in fetch_data.values() if samp == ""]
                if validator:
                    messagebox.showerror("ERROR", "Validation failed. Missing data")
                    return

                data["user_credentials"].append(fetch_data)
                with open("CREDS/user_creds.json", 'w') as upfile:
                    json.dump(data, upfile, indent=4)
                messagebox.showinfo("REGISTER", "SUCCESSFULLY REGISTERED.")
                new_win.destroy()
                root.deiconify()
            else:
                messagebox.showerror("ENGKK", "You are not an admin...")
                clear_widgets()
                return
        else:
            fetch_data = {
                "User_name": user_name.get(),
                "Password": encrypt_pw(password_entry_1.get()),
                "User_type": "Regular",
                "Profile_path": profile_pic_path.cget("text"),
                "Orders_History": [],
                "Total_Orders": 0,
                "is_logged_in": False
            }

            if password_entry_1.get() != password_entry_2.get():
                messagebox.showwarning("Read", "Your password doesn't match.")
                return
                
            if os.path.exists("CREDS/user_creds.json"):
                with open("CREDS/user_creds.json", 'r') as rfile:
                    data = json.load(rfile)
            else:
                data = {"user_credentials": []}
                with open("CREDS/user_creds.json", 'w') as wfile:
                    json.dump(data, wfile, indent=4)

            validator = [samp for samp in fetch_data.values() if samp == ""]
            if validator:
                messagebox.showerror("ERROR", "Validation failed. Missing data")
                return

            data["user_credentials"].append(fetch_data)
            with open("CREDS/user_creds.json", 'w') as upfile:
                json.dump(data, upfile, indent=4)
            messagebox.showinfo("REGISTER", "SUCCESSFULLY REGISTERED.")
            new_win.destroy()
            root.deiconify()
    
    
    
    def exit():
        if messagebox.askyesno("ALERT!", "Do you want to exit?"):
            new_win.withdraw()
            root.deiconify()
        else:
            return None

    new_win.protocol("WM_DELETE_WINDOW", exit)
    new_win.mainloop()

